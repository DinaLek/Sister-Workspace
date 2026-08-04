"""
מייצר pipeline.xlsx מתוך pipeline.csv
הרץ אחרי כל עדכון: python _crm/generate_excel.py
"""
import csv
import sys
import subprocess
import os

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl",
                           "--trusted-host", "pypi.org", "--trusted-host", "files.pythonhosted.org", "-q"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
CSV_PATH  = os.path.join(BASE_DIR, "pipeline.csv")
XLSX_PATH = os.path.join(BASE_DIR, "pipeline.xlsx")

HEADER_BG   = "ED1A47"   # Primary Pink של Sister (עדכון פלטה יולי 2026)
HEADER_TEXT = "FFFFFF"
ROW_ODD     = "FFFFFF"
ROW_EVEN    = "FAF8F8"   # לבן קל מאוד

STATUS_OPTIONS = "ליד חם,נשלחה הצעה,ממתין לתגובה,במשא ומתן,אושר,נדחה,נדחה לעתיד"

COLUMN_WIDTHS = [25, 16, 14, 24, 16, 13, 13, 35, 12, 16, 22, 20, 14, 35]

THIN_BORDER = Border(
    bottom=Side(style="thin", color="E0E0E0"),
)

def run():
    with open(CSV_PATH, encoding="utf-8") as f:
        rows = list(csv.reader(f))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pipeline"
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A2"

    header_font = Font(name="Arial", bold=True, color=HEADER_TEXT, size=11)
    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right",  vertical="center", wrap_text=True)
    data_font   = Font(name="Arial", size=10)

    # כותרות
    for col_idx, val in enumerate(rows[0], 1):
        cell = ws.cell(row=1, column=col_idx, value=val)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
    ws.row_dimensions[1].height = 32

    # נתונים
    for row_idx, row in enumerate(rows[1:], 2):
        bg = ROW_ODD if row_idx % 2 == 0 else ROW_EVEN
        row_fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")

        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = data_font
            cell.fill      = row_fill
            cell.alignment = right_align
            cell.border    = THIN_BORDER

            # עמודת סכום — מספר עם פורמט ₪
            if col_idx == 9 and val:
                try:
                    cell.value         = int(val)
                    cell.number_format = '#,##0'
                    cell.alignment     = Alignment(horizontal="center", vertical="center")
                except ValueError:
                    pass

            # עמודת סטטוס — מרכז
            if col_idx == 12:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[row_idx].height = 22

    # רוחב עמודות
    for i, width in enumerate(COLUMN_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Dropdown validation לעמודת סטטוס (L)
    dv = DataValidation(
        type="list",
        formula1=f'"{STATUS_OPTIONS}"',
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="סטטוס לא תקין",
        error="בחרי ערך מהרשימה"
    )
    ws.add_data_validation(dv)
    dv.sqref = "L2:L1000"

    wb.save(XLSX_PATH)
    print(f"OK: pipeline.xlsx updated ({len(rows)-1} leads)")

if __name__ == "__main__":
    run()
